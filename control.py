import openpyxl
from os import walk
import sys
import os

from getpass import getuser
import zipfile

import shutil
import openpyxl
from datetime import date
from datetime import datetime
import win32com.client as win32
from os import remove    
import pythoncom

def control (numerodesesion,id_cliente,token,user,alcance,facturas,entregas,factura_papel,pedido_scienza,
                                          pedido_cliente,remito_papel,clasedearmado,requierepresupuesto,requiereoc,fila,wb,ws,path,pathborrador):
    
    #pythoncom.CoInitialize()
    #path=('C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + str(id_cliente) + '.xlsm')
    #wb = openpyxl.load_workbook(path)
    #ws = wb["Iterador"]
    
    for elemento in range(0,alcance):
        if clasedearmado=="osde":
            archivoremitocopresupuesto="0" + str(factura_papel[elemento]) + "_" + str(remito_papel[elemento]) + '.pdf'
            archivofactura='0' + factura_papel[elemento] + '.pdf'
            pathoc='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + str(factura_papel[elemento]) + '/' + 'OC' + '/' + archivoremitocopresupuesto
            pathpresupuesto='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + 'Presupuestos' + '/' + archivoremitocopresupuesto
            pathfactura='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + "Facturas"  + '/' + archivofactura
            pathremito='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + 'Remitos'  + '/' + archivoremitocopresupuesto
        else:
            archivoremitocopresupuesto=str(remito_papel[elemento]) + ' - ' + str(pedido_cliente[elemento]) + '.pdf'
            archivofactura=factura_papel[elemento] + '.pdf'
            pathoc='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + str(factura_papel[elemento]) + '/' + 'OC' + '/' + archivoremitocopresupuesto
            pathpresupuesto='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + str(factura_papel[elemento]) + '/' + 'Presupuestos' + '/' + archivoremitocopresupuesto
            pathfactura='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + str(factura_papel[elemento])  + '/' + archivofactura
            pathremito='C:/Users/' + user +  '/Desktop/legajos' + '/' + str(id_cliente) + '/' + str(factura_papel[elemento]) + '/' + 'Remitos'  + '/' + archivoremitocopresupuesto
            
        if requiereoc[elemento]=="sin_oc":
            ws.cell(row=fila[elemento],column=16).value="sin_oc"
        else:
            if os.path.isfile(pathoc):
                ws.cell(row=int(fila[elemento]),column=16).value="esta la oc"
            else:
                ws.cell(row=int(fila[elemento]),column=16).value="no esta la oc"
    
        if requierepresupuesto[elemento]=="sinpresupuesto":
            ws.cell(row=int(fila[elemento]),column=15).value="sinpresupuesto"
        else:
            if os.path.isfile(pathpresupuesto):
                ws.cell(row=int(fila[elemento]),column=15).value="conpresupuesto"
            else:
                ws.cell(row=int(fila[elemento]),column=15).value="sinpresupuesto"
                
        if entregas[elemento]==None:
                ws.cell(row=int(fila[elemento]),column=13).value="descargaderemitook"
        else:
            if os.path.isfile(pathremito):
                ws.cell(row=int(fila[elemento]),column=13).value="descargaderemitook"
            else:
                ws.cell(row=int(fila[elemento]),column=13).value="descargaderemitonook"
        
        if os.path.isfile(pathfactura):
            ws.cell(row=int(fila[elemento]),column=14).value="descargadefacturaok"
        else:
            ws.cell(row=int(fila[elemento]),column=14).value="descargadefacturanook"
        
    wb.save(path)
    wb.close()
    os.remove(pathborrador)

