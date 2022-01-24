from zipfile import LargeZipFile
from autenticarse import loguearse
from descarga_remito import remitos
import openpyxl
import pythoncom
from getpass import getuser
from remito import proceso
import time
import os
from control import control
from datetime import date
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from tkinter import *
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def pdf_cap():
#-------------------------------------------------Ventana Raiz------------------------------------------------------#
    user=getuser()
    root=Tk()
    dt = datetime.now()    # Fecha y hora actual
    day = dt.day
    month = dt.month
    year = dt.year
    hoy=str(day) + "-" + str(month) + "-" + str(year)

    token=loguearse()
#-------------------------------------------------Interfaz Gráfica---------------------------------------------------------------#



#-------------------------------------------------Funciones-------------------------------------------------------------#
    #leer excel
    
    pythoncom.CoInitialize()
    path=('C:/Users/' + user +  '/Desktop/pdf_para_cap/liquidacion.xlsx')
    pathpdf='U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_PDF/'
    pathexcel='U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_EXCEL/'
    
    wb = openpyxl.load_workbook(path,keep_vba=False)
    ws = wb["CaP"]
    cuits=[]
    entregas=[]
    fechas=[]
    ultimafiladelws=len(ws['A'])

    #armar carpeta de cada legajo
    
    for dato in range(2,ultimafiladelws+1):
        if dato==None:
            continue
        else:
            cuits= ws.cell(row=dato,column=3).value
            entregas= ws.cell(row=dato,column=1).value
            fechas= ws.cell(row=dato,column=7).value
            # fecha= datetime.strftime("%YYYY%m%d")
            f="21-12-2021"
            pathcuits= 'U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_PDF/' + str(cuits)            
            pathfechas= 'U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_PDF/' + str(cuits)+ '/' + str(f)
            
            if os.path.isdir(pathcuits):
                if not os.path.isdir(pathfechas):
                    os.mkdir(pathfechas)
            else:
                os.mkdir(pathcuits)
                os.mkdir(pathfechas)
            
            
#---------------------------------------------------------fin-----------------------------------------------------------#
if __name__=="__main__":
    pdf_cap()    