from os import walk
import sys
import os
from getpass import getuser
import shutil
import openpyxl
from datetime import date
from datetime import datetime
import win32com.client as win32
from os import remove    


def tomardatos(wb,ws):
        factura_sap=[]
        entrega=[]
        pedido_scienza=[]
        pedido_cliente=[]
        remito=[]
        factura_papel=[]

        ultimafiladelws=len(ws['A'])

        for dato in range(2,ultimafiladelws+1):
            factura_sap.append(ws.cell(row=dato,column=1).value)
            entrega.append(ws.cell(row=dato,column=2).value)
            pedido_scienza.append(ws.cell(row=dato,column=3).value)
            pedido_cliente.append(ws.cell(row=dato,column=4).value)
            remito.append(ws.cell(row=dato,column=5).value)
            factura_papel.append(ws.cell(row=dato,column=6).value)
            
        return factura_sap,entrega,pedido_scienza,pedido_cliente,remito,factura_papel
